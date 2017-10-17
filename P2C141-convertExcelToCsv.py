#! python

import openpyxl, csv, os, sys

EXCEL_FOLDER = "P2C14/Excel Files"
CSV_FOLDER = "Converted_CSV"

os.makedirs(EXCEL_FOLDER, exist_ok=True)
os.chdir(EXCEL_FOLDER)

os.makedirs(CSV_FOLDER, exist_ok=True)

print("Searching for excel files in %s directory...." % EXCEL_FOLDER)
for file in os.listdir("."):
    if not file.endswith(".xlsx"):
        continue

    workbook = openpyxl.load_workbook(file)
    sheets = workbook.get_sheet_names()

    for sheet in sheets:
        sheetObj = workbook.get_sheet_by_name(sheet)

        csvFileName = os.path.join(CSV_FOLDER, file[:-5] + "_" + sheetObj.title + ".csv")
        csvFile = open(csvFileName,"w")
        csvWriter = csv.writer(csvFile)

        print("Creating csv for file %s" % csvFileName)

        for row in range(1, sheetObj.max_row + 1):
            rowData = []
            for col in range(1, sheetObj.max_column + 1):
                rowData.append(sheetObj.cell(row=row, column=col).value)

            csvWriter.writerow(rowData)

        csvFile.close()

print("Done")
#! python

import openpyxl, os, sys

if len(sys.argv) == 4:
    beforeRow = int(sys.argv[1])
    blankLines = int(sys.argv[2])
    fileName = sys.argv[3]
else:
    print("Run the program as P2C122-blankRowInserter.py [before row number] [number of blank lines] [filename from P2C12 folder]")
    sys.exit()

FOLDER_NAME = "./P2C12/"

os.makedirs(FOLDER_NAME, exist_ok=True)
os.chdir(FOLDER_NAME)

print("Opening file %s...." % fileName)
workbook = openpyxl.load_workbook(fileName)
sheet = workbook.get_active_sheet()

newWorkbook = openpyxl.Workbook()
newWorkbook.remove_sheet(newWorkbook.get_active_sheet())
newSheet = newWorkbook.create_sheet(sheet.title)

print("Reading and inserting blank lines....")
for row in range(1, beforeRow):
    for col in range(1, sheet.max_column+1):
        newSheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value

for row in range(beforeRow, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        newSheet.cell(row=row+blankLines, column=col).value = sheet.cell(row=row, column=col).value

print("Saving updated file....")
newWorkbook.save("updated_%s" % fileName)
workbook.close()
newWorkbook.close()

print("Done")


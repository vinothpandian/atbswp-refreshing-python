#! python

import openpyxl, os

FOLDER_NAME = "./P2C12/"
FILE_NAME = "example.xlsx"

os.makedirs(FOLDER_NAME, exist_ok=True)
os.chdir(FOLDER_NAME)

print("Opening file %s...." % FILE_NAME)
workbook = openpyxl.load_workbook(FILE_NAME)
newWorkbook = openpyxl.Workbook()

sheet = workbook.get_active_sheet()
newWorkbook.remove_sheet(newWorkbook.get_active_sheet())
newSheet = newWorkbook.create_sheet(sheet.title)

print("Reading and inverting data into new file....")
for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        newSheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

print("Saving the new file....")
newWorkbook.save("inverted_%s" % FILE_NAME)
workbook.close()
newWorkbook.close()

print("Done")
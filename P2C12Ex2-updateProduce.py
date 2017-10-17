#! python

import openpyxl,os

FOLDER_NAME = "./P2C12/"

os.makedirs(FOLDER_NAME, exist_ok=True)
os.chdir(FOLDER_NAME)

PRICE_UPDATES = {
    "Garlic": 3.07,
    "Celery": 1.19,
    "Lemon": 1.27
}

print("Opening workbook....")
workbook = openpyxl.load_workbook("produceSales.xlsx")
sheet = workbook.get_sheet_by_name("Sheet")

print("Reading and processing....")
for i in range(2, sheet.max_row +1):
    produceName = sheet.cell(row=i, column=1).value

    if produceName in PRICE_UPDATES:
        sheet.cell(row=i, column=2).value = PRICE_UPDATES[produceName]

print("Updated and saving....")
workbook.save("updatedProduceSales.xlsx")

print("Done")
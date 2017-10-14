#! python

import openpyxl, os, sys
from openpyxl.styles import Font

if len(sys.argv) == 2:
    number = int(sys.argv[1])
else:
    print("Run the program as P2C121-multiplicationTable.py [number]")
    sys.exit()

FOLDER_NAME = "./P2C12/"

os.makedirs(FOLDER_NAME, exist_ok=True)
os.chdir(FOLDER_NAME)

bold = Font(bold=True)

print("Creating workbook....")
workbook = openpyxl.Workbook()
sheet = workbook.get_active_sheet()

print("Calculating multiplication table....")
for row in range(1, number+2):
    for col in range(1, number+2):
        if row == 1:
            if col != 1:
                sheet.cell(row=1, column=col).value = col-1
                sheet.cell(row=1, column=col).font = bold
        else:
            if col == 1:
                sheet.cell(row=row, column=1).value = row-1
                sheet.cell(row=row, column=1).font = bold
            else:
                sheet.cell(row=row, column=col).value = (row-1)*(col-1)

print("Saving the file....")
workbook.save("multiplicationTableOf"+str(number)+".xlsx")
workbook.close()

print("Done")
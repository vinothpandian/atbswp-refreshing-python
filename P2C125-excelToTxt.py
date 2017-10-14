#! python

import openpyxl, os

FOLDER_NAME = "./P2C12/"
TEXTFILE_FOLDER = "convertedTextFiles"
FILE_NAME = "inverted_example.xlsx"

os.makedirs(FOLDER_NAME, exist_ok=True)
os.chdir(FOLDER_NAME)

os.makedirs(TEXTFILE_FOLDER , exist_ok=True)

print("Opening workbook %s...." % FILE_NAME)
workbook = openpyxl.load_workbook(FILE_NAME)
sheet = workbook.get_active_sheet()

print("Reading the workbook and creating text files....")
for col in range(1, sheet.max_column+1):
    txtFile = open(os.path.join(TEXTFILE_FOLDER, os.path.splitext(FILE_NAME)[0] + "_" + str(col) + ".txt"), 'w')
    for row in range(1, sheet.max_row+1):
        txtFile.write(str(sheet.cell(row=row, column=col).value)+"\n")
    print("Created text file from Column %d" % col)
    txtFile.close()

workbook.close()

print("Done")